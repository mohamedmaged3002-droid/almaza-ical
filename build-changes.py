#!/usr/bin/env python3
"""Build the "changed units only" xlsx from out/changed-units.json (written by
pricewatch.js). One row per changed date-range. Prices are EGP and ALREADY INCLUDE
BlueKeys' 10% markup (round(egp * 1.10)) — the SAME markup the OTA sheet bakes in
(build-sheet.py MARKUP, D-036) so the numbers line up with what the team lists.
old -> new + change. Attached to the change email by send-alert.js. No file is
written (and any stale one is removed) when nothing changed.

Mirrors brassbell-ical/build-changes.py, adapted to EGP + the 10% markup.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "out/changed-units.json"
OUT = "out/almaza-changes.xlsx"
MARKUP = 1.10   # keep in step with build-sheet.py MARKUP (D-036)


def marked(egp):
    return round(egp * MARKUP) if isinstance(egp, (int, float)) else None


def main():
    if not os.path.exists(SRC):
        print("build-changes: no changed-units.json — nothing to build.")
        return
    data = json.load(open(SRC))
    if not data or not data.get("units"):
        print("build-changes: no changes — not writing a sheet.")
        if os.path.exists(OUT):
            os.remove(OUT)
        return

    date_str = data.get("dateStr", "")
    cols = [
        ("Unit ID", 12), ("Unit", 40),
        ("From", 13), ("To", 13),
        ("Old (EGP incl 10%)", 18), ("New (EGP incl 10%)", 18), ("Change (EGP)", 14),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Changes"
    ws.append([c[0] for c in cols])
    for i, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    rows = []
    for u in data["units"]:
        for r in u.get("ranges", []):
            old = marked(r.get("oldEgp"))
            new = marked(r.get("newEgp"))
            change = (new - old) if isinstance(old, int) and isinstance(new, int) else None
            rows.append([u.get("wp"), u.get("title") or u.get("code") or u.get("wp"),
                         r.get("from"), r.get("to"), old, new, change])
    rows.sort(key=lambda x: (str(x[0]), str(x[2])))   # by wp then From
    for row in rows:
        ws.append(row)

    # styling (mirrors brassbell-ical/build-changes.py)
    hdr_fill = PatternFill("solid", fgColor="FF1F3B57")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    base = Font(name="Arial", size=10)
    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 26
    money_cols = (5, 6, 7)
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(ri, ci)
            cell.font = base
            cell.border = border
            if ci in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;-#,##0"
        chg = ws.cell(ri, 7)
        if isinstance(chg.value, (int, float)):
            chg.font = Font(name="Arial", size=10, bold=True,
                            color="FFB00020" if chg.value < 0 else "FF1B7A3D")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}{ws.max_row}"

    wb.save(OUT)
    print(f"build-changes: wrote {OUT} ({len(rows)} changed date-ranges across {len(data['units'])} units, {date_str}).")


if __name__ == "__main__":
    main()
