#!/usr/bin/env python3
"""build-sheet.py — OTA listing-build pack for Almaza Bay, in the Brassbell format.

Three tabs (mirrors "Brassbell Onboarding OTAs"):
  1. Almaza Master  — one row per unit (identity, photos, iCal, guests, coords).
  2. Monthly Prices — one row per unit; a column per month with the real nightly
     EGP. If a month's price changes mid-month the cell shows the exact day
     ranges (e.g. "1–22: 19,000 / 23–31: 22,000"). Per-row green→red heatmap.
  3. Price Ranges   — one row per continuous date range at one flat nightly EGP.

Prices are the operator's real EGP rates (named period price, else the Default
Rate) — exactly as shown on almazabay.lodgify.com. No currency conversion. The
OTA team converts at listing time. NO network, NO DB.
"""
import json
import os
import re
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
UNITS_DIR = os.path.join(HERE, "output", "units")
INDEX_JSON = os.path.join(HERE, "docs", "index.json")
DAILY_JSON = os.path.join(HERE, "output", "daily-prices.json")
OUT = os.path.join(HERE, "Almaza Master.xlsx")

CUR = "EGP"
ICAL_BASE = "https://mohamedmaged3002-droid.github.io/almaza-ical/"
GALLERY_BASE = "https://mohamedmaged3002-droid.github.io/almaza-ical/photos/"

# Almaza's operator rates only cover the summer season Jun–Oct 2026.
MONTHS = [("06", "Jun '26"), ("07", "Jul '26"), ("08", "Aug '26"),
          ("09", "Sep '26"), ("10", "Oct '26")]

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
BLOCKED_FILL = PatternFill("solid", fgColor="D9D9D9")   # grey = unavailable
DOCS_DIR = os.path.join(HERE, "docs")

SEASON_START = date(2026, 6, 1)
SEASON_END = date(2026, 10, 31)


def season_dates():
    out, d = [], SEASON_START
    while d <= SEASON_END:
        out.append(d.isoformat())
        d += timedelta(days=1)
    return out


def load_blocked(wp):
    """Set of ISO dates this unit is BLOCKED, parsed from its docs/{wp}.ics feed
    (DTSTART inclusive .. DTEND exclusive). Point-in-time snapshot of the feed."""
    p = os.path.join(DOCS_DIR, f"{wp}.ics")
    blocked = set()
    if not os.path.exists(p):
        return blocked
    with open(p, encoding="utf-8") as f:
        txt = f.read()
    for ev in txt.split("BEGIN:VEVENT")[1:]:
        ds = re.search(r"DTSTART;VALUE=DATE:(\d{4})(\d{2})(\d{2})", ev)
        de = re.search(r"DTEND;VALUE=DATE:(\d{4})(\d{2})(\d{2})", ev)
        if not ds or not de:
            continue
        d = date(*map(int, ds.groups()))
        end = date(*map(int, de.groups()))          # exclusive
        while d < end:
            blocked.add(d.isoformat())
            d += timedelta(days=1)
    return blocked


# ----- data loading ----------------------------------------------------------
CONTENT_BUNDLE = os.path.join(HERE, "data", "sheet-content.json")


def load_units():
    # Prefer the committed static bundle (so CI can build without output/units/*),
    # fall back to the local scrape output.
    if os.path.exists(CONTENT_BUNDLE):
        with open(CONTENT_BUNDLE, encoding="utf-8") as f:
            units = json.load(f)
    else:
        units = []
        for name in os.listdir(UNITS_DIR):
            if name.endswith(".json"):
                with open(os.path.join(UNITS_DIR, name), encoding="utf-8") as f:
                    units.append(json.load(f))
    units.sort(key=lambda u: u.get("wp", 0))
    return units


def photo_count(u):
    return u["photoCount"] if "photoCount" in u else len(u.get("photos") or [])


def load_min_stays():
    if not os.path.exists(INDEX_JSON):
        return {}
    try:
        with open(INDEX_JSON, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}
    return {p["wp"]: p.get("minStay") for p in data.get("properties", []) if p.get("wp") is not None}


def load_daily():
    if not os.path.exists(DAILY_JSON):
        return {}
    with open(DAILY_JSON, encoding="utf-8") as f:
        return json.load(f)


# BlueKeys 10% markup baked into the OTA-sheet nightly prices (D-036), so the OTA
# team lists the final guest price directly — no mental math. On bluekeys.co the
# same 10% shows as a separate "Service fee (10%)" line; unit_daily_prices in the
# DB stays RAW (this markup is sheet-presentation only).
MARKUP = 1.10


def marked(egp):
    return round(egp * MARKUP)


def daily_of(daily, u):
    rows = daily.get(str(u.get("wp"))) or daily.get(u.get("wp")) or []
    return [{"date": r["date"], "price": marked(r["price"])} for r in rows]  # +10%


# ----- pricing helpers (EGP, incl. 10% markup) -------------------------------
def money(egp):
    return f"{egp:,}"                     # e.g. 20900 -> "20,900"


def month_segments(daily_rows, mm):
    """Day-of-month runs of one price in month mm: [{d1, d2, egp}]."""
    days = sorted((r for r in daily_rows if r["date"][5:7] == mm), key=lambda x: x["date"])
    segs = []
    for r in days:
        d, p = int(r["date"][8:10]), r["price"]
        if segs and segs[-1]["egp"] == p and segs[-1]["d2"] == d - 1:
            segs[-1]["d2"] = d
        else:
            segs.append({"d1": d, "d2": d, "egp": p})
    return segs


def month_cell(daily_rows, mm):
    """(display_text, representative_egp_for_colour). Blank if no priced days."""
    segs = month_segments(daily_rows, mm)
    if not segs:
        return "", None
    if len(segs) == 1:
        return money(segs[0]["egp"]), segs[0]["egp"]
    lines = [(f"{s['d1']}–{s['d2']}: {money(s['egp'])}" if s["d1"] != s["d2"] else f"{s['d1']}: {money(s['egp'])}")
             for s in segs]
    prices = sorted(r["price"] for r in daily_rows if r["date"][5:7] == mm)
    return "\n".join(lines), prices[len(prices) // 2]   # median for the heatmap


def heat_color(t):
    """green #63BE7B -> yellow #FFEB84 -> red #F8696B, t in [0,1]."""
    g, y, r = (0x63, 0xBE, 0x7B), (0xFF, 0xEB, 0x84), (0xF8, 0x69, 0x6B)
    a, b, tt = (g, y, t * 2) if t <= 0.5 else (y, r, (t - 0.5) * 2)
    c = tuple(round(a[i] + (b[i] - a[i]) * tt) for i in range(3))
    return f"{c[0]:02X}{c[1]:02X}{c[2]:02X}"


def next_day(iso_date):
    y, m, d = map(int, iso_date.split("-"))
    return (date(y, m, d) + timedelta(days=1)).isoformat()


def segments_for(daily_rows):
    """Full-season contiguous same-price runs: [{start, end, nights, egp}]."""
    segs = []
    for r in sorted(daily_rows, key=lambda x: x["date"]):
        p = r["price"]
        if segs and segs[-1]["egp"] == p and next_day(segs[-1]["end"]) == r["date"]:
            segs[-1]["end"] = r["date"]
            segs[-1]["nights"] += 1
        else:
            segs.append({"start": r["date"], "end": r["date"], "nights": 1, "egp": p})
    return segs


# ----- sheet builders --------------------------------------------------------
def style_headers(ws, row):
    for cell in ws[row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")


def build_master(ws, units, min_stays):
    window = eligibility_window()
    ws.append(["Almaza Bay — OTA listing pack"])
    ws.append([f"One row per unit. Prices (EGP) in the Monthly Prices / Price Ranges tabs ALREADY INCLUDE "
               f"BlueKeys' 10% markup — list them as-is (do NOT add anything). 'ota_eligible' = YES when the "
               f"unit has >= {ELIG_MIN} available nights (not blocked) between today and 1 Oct 2026; refreshes daily."])
    ws.append([])
    cols = ["wp_post_id", "source_code", "operator_unit_code", "sub_community", "title",
            "property_type", "guests_bluekeys", "guests_operator", "bedrooms", "bathrooms",
            "default_rate_egp", "min_stay", "checkin_time", "checkout_time",
            "amenities", "photo_gallery", "photo_count", "ical_url",
            "lat", "lng", "source_url", "avail_nights_to_1oct", "ota_eligible"]
    ws.append(cols)
    style_headers(ws, 4)
    elig_col = len(cols)                 # last column = ota_eligible
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    r, n_elig = 5, 0
    for u in units:
        rates = u.get("rates") or {}
        lat, lng = u.get("lat"), u.get("lng")
        avail = available_nights(u.get("wp"), window)
        elig = avail >= ELIG_MIN
        if elig:
            n_elig += 1
        ws.append([
            u.get("wp"), u.get("sourceCode"), u.get("operatorCode"),
            u.get("subCommunity") or "UNKNOWN — needs review", u.get("title"),
            "Vacation Rental", u.get("guestsBluekeys"), u.get("guestsOperator"),
            u.get("bedrooms"), u.get("bathrooms"),
            marked(rates.get("defaultRate")) if rates.get("defaultRate") is not None else "",
            min_stays.get(u.get("wp"), ""), u.get("checkinTime"), u.get("checkoutTime"),
            ", ".join(u.get("amenities") or []),
            GALLERY_BASE + str(u.get("wp")) + ".html", photo_count(u),
            ICAL_BASE + str(u.get("wp")) + ".ics",
            "NEEDS PIN" if lat is None else lat, "NEEDS PIN" if lng is None else lng,
            u.get("sourceUrl"), avail, "YES" if elig else "NO",
        ])
        ws.cell(r, elig_col).fill = green if elig else red
        r += 1
    widths = {"title": 42, "sub_community": 20, "amenities": 50, "photo_gallery": 56,
              "ical_url": 56, "source_url": 56, "checkin_time": 12, "checkout_time": 16,
              "default_rate_egp": 15, "avail_nights_to_1oct": 20, "ota_eligible": 13}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(4, i).column_letter].width = widths.get(c, 13)
    ws.freeze_panes = "A5"
    return n_elig


def build_monthly(ws, units, daily):
    ws.append(["Nightly price by month — EGP, incl. BlueKeys 10% markup (one row per listing)"])
    ws.append(['One row per unit. Each month = the nightly EGP to LIST ON THE OTA — it already includes '
               'BlueKeys’ 10% markup, so use it as-is (do not add anything). If a month splits '
               '(e.g. "1–22: 20,900 / 23–31: 24,200") the price changed mid-month, with the exact days. '
               'Operator season Jun–Oct 2026 only. Colour: green = low → red = peak (per row).'])
    ws.append([])
    cols = ["wp", "Code/Slug", "Title", "Area", "Beds"] + [lbl for _, lbl in MONTHS]
    ws.append(cols)
    style_headers(ws, 4)

    r = 5
    for u in units:
        rows = daily_of(daily, u)
        blocked = load_blocked(u.get("wp"))
        cells, reps, blkd_flags = [], [], []
        for mm, _ in MONTHS:
            mdates = [x["date"] for x in rows if x["date"][5:7] == mm]
            fully_blocked = bool(mdates) and all(d in blocked for d in mdates)
            if fully_blocked:
                cells.append("blkd")
                reps.append(None)
                blkd_flags.append(True)
            else:
                txt, rep = month_cell(rows, mm)
                cells.append(txt)
                reps.append(rep)
                blkd_flags.append(False)
        ws.append([u.get("wp"), u.get("slug"), u.get("title"),
                   u.get("subCommunity") or "", u.get("bedrooms")] + cells)
        valid = [x for x in reps if x is not None]
        lo, hi = (min(valid), max(valid)) if valid else (0, 0)
        max_lines = 1
        for j, rep in enumerate(reps):
            cell = ws.cell(r, 6 + j)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if blkd_flags[j]:
                cell.fill = BLOCKED_FILL
                continue
            if rep is None:
                continue
            t = (rep - lo) / (hi - lo) if hi > lo else 0.0
            cell.fill = PatternFill("solid", fgColor=heat_color(t))
            max_lines = max(max_lines, str(cell.value).count("\n") + 1)
        ws.row_dimensions[r].height = 15 * max_lines
        r += 1

    for i, w in enumerate([8, 26, 34, 18, 6] + [17] * len(MONTHS), 1):
        ws.column_dimensions[ws.cell(4, i).column_letter].width = w
    ws.freeze_panes = "F5"


def merged_segments(daily_rows, blocked):
    """Chronological segments over the season: each is either a flat-price run
    ('egp') or a Blocked run (unavailable per the iCal feed). Blocked wins over
    price on a date. [{start, end, nights, egp|None, blocked}]."""
    price_by = {r["date"]: r["price"] for r in daily_rows}
    segs = []
    for iso in season_dates():
        if iso in blocked:
            key = ("BLK", None)
        elif iso in price_by:
            key = ("EGP", price_by[iso])
        else:
            continue
        if segs and segs[-1]["key"] == key and next_day(segs[-1]["end"]) == iso:
            segs[-1]["end"] = iso
            segs[-1]["nights"] += 1
        else:
            segs.append({"start": iso, "end": iso, "nights": 1, "key": key})
    return segs


def build_ranges(ws, units, daily):
    ws.append(["Nightly price by date range — EGP, incl. BlueKeys 10% markup (exact, no estimation)"])
    ws.append(["Each row = a continuous date range at one flat nightly rate to LIST ON THE OTA (Almaza's "
               "rate + BlueKeys' 10% markup, already baked in — use as-is). Grey 'Blocked' rows = nights the "
               "calendar shows unavailable (snapshot of the live iCal feed — re-check the feed before booking)."])
    ws.append([])
    cols = ["wp", "Code/Slug", "Title", "Area", "Beds", "From", "To", "Nights", "Nightly EGP (incl. 10%)"]
    ws.append(cols)
    style_headers(ws, 4)
    n = n_blk = 0
    r = 5
    for u in units:
        blocked = load_blocked(u.get("wp"))
        for s in merged_segments(daily_of(daily, u), blocked):
            is_blk = s["key"][0] == "BLK"
            ws.append([u.get("wp"), u.get("slug"), u.get("title"), u.get("subCommunity") or "",
                       u.get("bedrooms"), s["start"], s["end"], s["nights"],
                       "Blocked" if is_blk else s["key"][1]])
            if is_blk:
                for c in ws[r]:
                    c.fill = BLOCKED_FILL
                n_blk += 1
            n += 1
            r += 1
    for i, w in enumerate([8, 26, 34, 18, 6, 13, 13, 8, 13], 1):
        ws.column_dimensions[ws.cell(4, i).column_letter].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{ws.max_row}"
    return n, n_blk


ELIG_END = date(2026, 10, 1)   # OTA-eligibility window end (1 Oct)
ELIG_MIN = 30                  # need >= 30 available nights in the window


def eligibility_window():
    """[today .. 1 Oct 2026] as ISO strings. Recomputed each build (moves daily)."""
    window, d = [], date.today()
    while d <= ELIG_END:
        window.append(d.isoformat())
        d += timedelta(days=1)
    return window


def available_nights(wp, window):
    """Nights in the window NOT blocked per the unit's live iCal feed."""
    blocked = load_blocked(wp)
    return sum(1 for iso in window if iso not in blocked)


def main():
    units = load_units()
    min_stays = load_min_stays()
    daily = load_daily()

    wb = Workbook()
    n_elig = build_master(wb.active, units, min_stays)
    wb.active.title = "Almaza Master"
    build_monthly(wb.create_sheet("Monthly Prices"), units, daily)
    n_seg, n_blk = build_ranges(wb.create_sheet("Price Ranges"), units, daily)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Tabs: {wb.sheetnames}")
    print(f"Units: {len(units)}  |  Price-range rows: {n_seg} ({n_blk} Blocked)  |  Currency: {CUR}")
    print(f"OTA-eligible (>= {ELIG_MIN} avail nights, today->1 Oct): {n_elig}/{len(units)}")


if __name__ == "__main__":
    main()
